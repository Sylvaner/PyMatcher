#!/usr/bin/env python3
"""
Match 2 columns in spreadsheet files and merge row in output file.
"""

__author__ = "Sylvain Dangin"
__licence__ = "Apache 2.0"
__version__ = "1.0"
__maintainer__ = "Sylvain Dangin"
__email__ = "sylvain.dangin@gmail.com"
__status__ = "Development"

import os
import sys
import openpyxl
import xlrd
import xlwt

class Matcher():
    output_sheetname = ''
    ignore_header = False
    match_column_1 = -1
    match_column_2 = -1
    input_data_1 = []
    input_data_2 = []

    def __init__(self):
        """Constructor of the class
        """
        self.set_default_values()

    def set_default_values(self):
        """Set default values of the class
        """
        self.output_sheetname = 'match'
        self.ignore_header = False
        self.match_column_1 = -1
        self.match_column_2 = -1
        self.input_data_1 = []
        self.input_data_2 = []

    def read_input_data_from_xls(self, input_filename):
        """Read input data in old Excel file (xls)

        :param input_filename: Path of the input file

        :return: Array with the data of the file.
        :rtype: Array
        """
        return_data = []

        if not os.path.exists(input_filename):
            raise Exception('Error: Input file not found.')

        workbook = xlrd.open_workbook(filename = input_filename)
        worksheet = workbook.sheet_by_index(0)
        
        for row_index, row_data in enumerate(range(worksheet.nrows)):
            row = []
            for column_index, column_data in enumerate(range(worksheet.ncols)):
                row.append(worksheet.cell_value(row_index, column_index))
            return_data.append(row)
        return return_data
        
    def write_output_data_to_xls(self, output_filename, output_data, output_sheetname = 'match'):
        """Write output data in old Excel file (xls)

        :param output_filename: Path of the output file
        :param output_data: Data to write.
        """
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(output_sheetname)

        for row_index, row in enumerate(output_data):
            for col_index, col in enumerate(row):
                worksheet.write(row_index, col_index, col)
        workbook.save(output_filename)

    def read_input_data_from_xlsx(self, input_filename):
        """Read input data in new Excel file (xlsx)

        :param input_filename: Path of the input file

        :return: Array with the data of the file.
        :rtype: Array
        """
        return_data = []

        if not os.path.exists(input_filename):
            raise Exception('Error: Input file not found.')

        workbook = openpyxl.load_workbook(input_filename)
        worksheet = workbook.worksheets[0]

        for row_index, row_data in enumerate(worksheet.rows):
            row = []
            for col_data in row_data:
                row.append(col_data.value)
            return_data.append(row)
        return return_data
    
    def write_output_data_to_xlsx(self, output_filename, output_data, output_sheetname = 'match'):
        """Write output data in new Excel file (xlsx)

        :param output_filename: Path of the output file
        :param output_data: Data to write.
       """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = output_sheetname

        for row in output_data:
            worksheet.append(row)
        workbook.save(output_filename)

    def read_argv(self, argv):
        """Read command line arguments and extract params
        """
        if len(argv) > 5:
            self.set_default_values()

            try:
                for arg in argv:
                    if '--' == arg[:2]:
                        if 'no-header' == arg[2:]:
                            self.ignore_header = True
                        elif 'output-sheetname' in arg:
                            self.output_sheetname = arg.split('=')[1]
                # Match column argument
                column = argv[len(argv) - 4]
                if column.isdigit():
                    self.match_column_1 = int(column)
                column = argv[len(argv) - 2]
                if column.isdigit():
                    self.match_column_2 = int(column)
                if self.match_column_1 < 1 or self.match_column_2 < 1:
                    return False, 'Bad column'

                return True, ''
            except Exception as e:
                return False, 'Error in arguments'
        return False, 'Not enough arguments'

    def match(self):
        """Match on 2 columns and merge data

        :return: Merged data
        :rtype: Array
        """
        output_data = []
        match_column_1 = self.match_column_1 - 1
        match_column_2 = self.match_column_2 - 1
        # Copy header
        if not self.ignore_header:
            row = []
            row.append(self.input_data_1[0][match_column_1])
            for index, col in enumerate(self.input_data_1[0]):
                if index != match_column_1:
                    row.append(col)
            for index, col in enumerate(self.input_data_2[0]):
                if index != match_column_2:
                    row.append(col)
            output_data.append(row)
        for index_1, input_1 in enumerate(self.input_data_1):
            if index_1 > 0:
                for index_2, input_2 in enumerate(self.input_data_2):
                    # Match columns
                    if index_2 > 0 and input_1[match_column_1] == input_2[match_column_2]:
                        # Merge the 2 inputs
                        row = []
                        # Matched data
                        row.append(input_1[match_column_1])
                        # Input 1 row data
                        for index, col in enumerate(input_1):
                            if index != match_column_1:
                                row.append(col)
                        # Input 2 row data
                        for index, col in enumerate(input_2):
                            if index != match_column_2:
                                row.append(col)
                        output_data.append(row)
        return output_data
    
    def start(self, input1_filename, input2_filename, output_filename):
        self.input_data_1 = []
        self.input_data_2 = []

        # Test if input files exists
        if not os.path.exists(input1_filename) or not os.path.exists(input2_filename):
            raise Exception('Error: input file missing')

        # Read data from the first file
        if '.xlsx' in input1_filename:
            self.input_data_1 = self.read_input_data_from_xlsx(input1_filename)
        elif '.xls' in input1_filename:
            self.input_data_1 = self.read_input_data_from_xls(input1_filename)

        # Read data from the second file
        if '.xlsx' in input2_filename:
            self.input_data_2 = self.read_input_data_from_xlsx(input2_filename)
        elif '.xls' in input2_filename:
            self.input_data_2 = self.read_input_data_from_xls(input2_filename)

        if self.input_data_1 == [] or self.input_data_2 == [] or self.input_data_1 == [[None]] or self.input_data_2 == [[None]]:
            raise Exception('Error: no data')

        # Start process
        output_data = self.match()
        if '.xlsx' in output_filename:
            self.write_output_data_to_xlsx(output_filename, output_data, self.output_sheetname)
        elif '.xls' in output_filename:
            self.write_output_data_to_xls(output_filename, output_data, self.output_sheetname)
    
def usage(exec_name):
    """Show usage for help.
    :param exec_name: Path of this script
    """

    # Get only executable name without the path
    if os.path.sep in exec_name:
        exec_name = exec_name.split(os.path.sep)[-1:][0]
        
    print(exec_name+" [options] input_file_1 match_column_1 input_file_2 match_column_2 output_file")
    print("Options : ")
    print(" --no-header")
    print(" --output-sheetname=SHEETNAME")
    
# Entry point
if __name__ == '__main__':
    argv = sys.argv
    matcher = Matcher()
    result, err = matcher.read_argv(argv)
    if result:
        nb_args = len(argv)
        matcher.start(argv[nb_args - 5], argv[nb_args - 3], argv[nb_args - 1])
    else:
        usage(argv[0])

