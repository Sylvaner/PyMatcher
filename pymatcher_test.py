#!/usr/bin/env python3
"""
Test pymatcher.Matcher class.
"""
__author__ = "Sylvain Dangin"
__licence__ = "Apache 2.0"
__version__ = "1.0"
__maintainer__ = "Sylvain Dangin"
__email__ = "sylvain.dangin@gmail.com"
__status__ = "Development"

import os
import unittest
import xlrd
import openpyxl
from pymatcher import Matcher

class MatcherTest(unittest.TestCase):
    # Base tests data
    TESTS_DATA_1 = [
            ['Activated', 'Money', 'Name', 'Birthdate'],
            ['Yes', '24', 'Marc ASSIN', '02/10/1970'],
            ['Yes', '89', 'John SMITH', '31/01/1988'],
            ['No', '0', 'Karl DO', '14/07/1972'],
            ['Yes', '69', 'Jean-Luc PASDIDEE', '06/07/1977']
        ]
    TESTS_DATA_2 = [
            ['City', 'Name', 'Favorite color'],
            ['New-York', 'Jean-Luc PASDIDEE', 'Green'],
            ['Troyes', 'John SMITH', 'Red'],
            ['Paris', 'Marc ASSIN', 'Purple'],
            ['London', 'Karl DO', 'Blue']
        ]
    # Directory where test files will be create
    TEST_DIRECTORY = 'tmp'
    # File for tests
    TEST_FILENAME = 'input'
    # Path of the first test file
    TEST_FILE_1_PATH = ''
    # Path of the second test file
    TEST_FILE_2_PATH = ''
    # Base name of the output test file
    OUTPUT_BASE_FILENAME = 'output'
    # Full base path of the output test file
    OUTPUT_BASE_FILE_PATH = ''
    # self.convert object
    matcher = None

    @classmethod
    def setUpClass(cls):
        """Set up configuration for tests
        """
        if not os.path.exists(cls.TEST_DIRECTORY):
            os.mkdir(cls.TEST_DIRECTORY)
        for file in os.listdir(cls.TEST_DIRECTORY):
            os.remove(cls.TEST_DIRECTORY+os.path.sep+file)
        cls.TEST_FILE_1_PATH = cls.TEST_DIRECTORY+os.path.sep+cls.TEST_FILENAME+'1.xlsx'
        cls.TEST_FILE_2_PATH = cls.TEST_DIRECTORY+os.path.sep+cls.TEST_FILENAME+'2.xlsx'
        cls.OUTPUT_BASE_FILE_PATH = cls.TEST_DIRECTORY+os.path.sep+cls.OUTPUT_BASE_FILENAME

    @classmethod
    def tearDownClass(cls):
        """Remove test directory at the end of tests
        """
        os.rmdir(cls.TEST_DIRECTORY)

    def setUp(self):
        """Initialise self.matcher class and create the test file
        """
        self.matcher = Matcher()
        self.create_test_file(self.TESTS_DATA_1, self.TEST_FILE_1_PATH)
        self.create_test_file(self.TESTS_DATA_2, self.TEST_FILE_2_PATH)
        
    def tearDown(self):
        """Remove all files after each test
        """
        for file in os.listdir(self.TEST_DIRECTORY):
            os.remove(self.TEST_DIRECTORY+os.path.sep+file)

    def get_cell_in_xls(self, filename, row, col, sheet = None):
        """Get cell value of old Excel file

        :param row: Row of the cell
        :param col: Column of the cell
        :param sheet: Sheet in Excel workbook
        """
        workbook = xlrd.open_workbook(filename = filename)
        worksheet = None
        if sheet is None:
            worksheet = workbook.sheet_by_index(0)
        else:
            worksheet = workbook.sheet_by_name(sheet)
        return worksheet.cell_value(row - 1, col - 1)

    def get_cell_in_xlsx(self, filename, row, col, sheet = None):
        """Get cell value of new Excel file

        :param row: Row of the cell
        :param col: Column of the cell
        :param sheet: Sheet in Excel workbook
        """
        workbook = openpyxl.load_workbook(filename)
        worksheet = None
        if sheet is None:
            worksheet = workbook.worksheets[0]
        else:
            worksheet = workbook.get_sheet_by_name(sheet)
        return worksheet.cell(row = row, column = col).value

    def create_test_file(self, data, filename):
        """Write output data in new Excel file (xlsx)
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Test"

        for row in data:
            worksheet.append(row)
        workbook.save(filename)

################################################################################
# Tests for file functions                                                     #
################################################################################
    def test_create_test_file(self):
        """Test creation of test file
        """
        self.assertEqual(self.TESTS_DATA_1[0][2], self.get_cell_in_xlsx(self.TEST_FILE_1_PATH, 1, 3))
        self.assertEqual(self.TESTS_DATA_2[3][1], self.get_cell_in_xlsx(self.TEST_FILE_2_PATH, 4, 2))

    def test_read_input_data_from_xlsx(self):
        """Test read xlsx method
        """
        data = self.matcher.read_input_data_from_xlsx(self.TEST_FILE_1_PATH)
        self.assertEqual(self.TESTS_DATA_1[1][2], data[1][2])

    def test_write_output_data_to_xlsx(self):
        """Test write xlsx method
        """
        self.matcher.write_output_data_to_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xls', self.TESTS_DATA_1)
        self.assertEqual(self.TESTS_DATA_1[1][2], self.get_cell_in_xls(self.OUTPUT_BASE_FILE_PATH+'.xls', 2, 3))

    def test_write_output_data_to_xlsx_with_sheet(self):
        """Test write xlsx method with specified sheetname
        """
        self.matcher.write_output_data_to_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xls', self.TESTS_DATA_1, 'just_a_sheet')
        self.assertEqual(self.TESTS_DATA_1[1][2], self.get_cell_in_xls(self.OUTPUT_BASE_FILE_PATH+'.xls', 2, 3, 'just_a_sheet'))

    def test_read_write_xls(self):
        """Test XLS methods
        """
        data = self.matcher.read_input_data_from_xlsx(self.TEST_FILE_2_PATH)
        self.matcher.write_output_data_to_xls(self.OUTPUT_BASE_FILE_PATH+'.xls', data)
        read_data = self.matcher.read_input_data_from_xls(self.OUTPUT_BASE_FILE_PATH+'.xls')
        self.assertEqual(self.TESTS_DATA_2[2][2], read_data[2][2])

################################################################################
# Tests for args functions                                                     #
################################################################################
    def test_no_enough_arguments(self):
        """Test without enough arguments
        """
        argv = ['exec', 'coucou']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(False, test)
        
        argv = ['exec', 'input1', 'input2']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(False, test)

    def test_with_bad_column_argument(self):
        """Test with bad column argument
        """
        argv = ['exec', 'input1', 'A', 'input2', '1', 'output']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(False, test)
        argv = ['exec', 'input1', '1', 'input2', 'B', 'output']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(False, test)

    def test_with_zero_column_argument(self):
        """Test with 0 column argument
        """
        argv = ['exec', 'input1', '0', 'input2', '1', 'output']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(False, test)
        argv = ['exec', 'input1', '1', 'input2', '0', 'output']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(False, test)

    def test_column_argument(self):
        """Test column argument
        """
        argv = ['exec', 'input1', '1', 'input2', '2', 'output']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(True, test)
        self.assertEqual(1, self.matcher.match_column_1)
        self.assertEqual(2, self.matcher.match_column_2)

    def test_no_params_in_arguments(self):
        """Test with only necessary arguments
        """
        argv = ['exec', 'input1', '1', 'input2', '3', 'output']
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(True, test)
        self.assertEqual(False, self.matcher.ignore_header)
        self.assertEqual('match', self.matcher.output_sheetname)
        self.assertEqual(1, self.matcher.match_column_1)
        self.assertEqual(3, self.matcher.match_column_2)

    def test_read_no_header_argument(self):
        """Test with no header arguments
        """
        argv = ['exec', '--no-header', 'input1', '1', 'input2', '2', 'output']
        self.assertEqual(False, self.matcher.ignore_header)
        test, err = self.matcher.read_argv(argv)
        self.assertEqual(True, test)
        self.assertEqual(True, self.matcher.ignore_header)

    def test_read_output_sheetname_argument(self):
        """Test with output-sheetname argument
        """
        argv = ['exec', '--output-sheetname=just_a_name', 'input1', '1', 'input2', '3', 'output']
        # Before
        self.assertEqual('match', self.matcher.output_sheetname)
        
        test, err = self.matcher.read_argv(argv)
        # After
        self.assertEqual(True, test)
        self.assertEqual('just_a_name', self.matcher.output_sheetname)

    def test_read_multiple_arguments(self):
        """Test with multiple arguments
        """
        argv = ['exec', '--no-header', '--output-sheetname=just_a_name', 'input1', '1', 'input2', '2', 'output']
        # Before
        self.assertEqual(False, self.matcher.ignore_header)
        self.assertEqual('match', self.matcher.output_sheetname)
        
        test, err = self.matcher.read_argv(argv)
        # After
        self.assertEqual(True, test)
        self.assertEqual(True, self.matcher.ignore_header)
        self.assertEqual('just_a_name', self.matcher.output_sheetname)

################################################################################
# Tests for merge function                                                     #
################################################################################
    def test_without_matches(self):
        """Test 2 array without matches
        """
        self.matcher.match_column_1 = 1
        self.matcher.match_column_2 = 1
        self.matcher.input_data_1 = self.TESTS_DATA_1
        self.matcher.input_data_2 = self.TESTS_DATA_2
        result = self.matcher.match()
        self.assertEqual('Activated', result[0][0])
        self.assertEqual('Money', result[0][1])
        self.assertEqual('Name', result[0][4])
        self.assertEqual(1, len(result))

    def test_with_matches(self):
        """Test 2 array without matches
        """
        self.matcher.match_column_1 = 3
        self.matcher.match_column_2 = 2
        self.matcher.input_data_1 = self.TESTS_DATA_1
        self.matcher.input_data_2 = self.TESTS_DATA_2
        result = self.matcher.match()
        self.assertEqual('Name', result[0][0])
        self.assertEqual('Activated', result[0][1])
        self.assertEqual('Favorite color', result[0][5])
        self.assertEqual('Marc ASSIN', result[1][0])
        self.assertEqual('London', result[3][4])

################################################################################
# Tests for start function                                                     #
################################################################################
    def test_no_input1_file(self):
        """Test if the first input file doesn't exists
        """
        with self.assertRaises(Exception):
            self.matcher.start(self.TEST_FILE_1_PATH, 'a', self.OUTPUT_BASE_FILE_PATH+'.xlsx')

    def test_no_input2_file(self):
        """Test if the second input file doesn't exists
        """
        with self.assertRaises(Exception):
            self.matcher.start('a', self.TEST_FILE_2_PATH, self.OUTPUT_BASE_FILE_PATH+'.xlsx')

    def test_no_input_file(self):
        """Test if all input files doesn't exists
        """
        with self.assertRaises(Exception):
            self.matcher.start('a', 'b', self.OUTPUT_BASE_FILE_PATH+'.xlsx')

    def test_empty_input1_file(self):
        """Test with first file empty
        """
        self.matcher.write_output_data_to_xlsx(self.TEST_FILE_1_PATH, [])
        with self.assertRaises(Exception):
            self.matcher.start(self.TEST_FILE_1_PATH, self.TEST_FILE_2_PATH, self.OUTPUT_BASE_FILE_PATH+'.xlsx')

    def test_empty_input2_file(self):
        """Test with second file empty
        """
        self.matcher.write_output_data_to_xls(self.TEST_FILE_2_PATH, [])
        with self.assertRaises(Exception):
            self.matcher.start(self.TEST_FILE_1_PATH, self.TEST_FILE_2_PATH, self.OUTPUT_BASE_FILE_PATH+'.xlsx')

    def test_two_xls_input_files(self):
        """Test readed data from two old Excel files
        """
        TEST_FILE_1_PATH = self.TEST_DIRECTORY+os.path.sep+self.TEST_FILENAME+'1.xls'
        TEST_FILE_2_PATH = self.TEST_DIRECTORY+os.path.sep+self.TEST_FILENAME+'2.xls'
        self.create_test_file(self.TESTS_DATA_1, TEST_FILE_1_PATH)
        self.create_test_file(self.TESTS_DATA_2, TEST_FILE_2_PATH)
        self.matcher.start(TEST_FILE_1_PATH, TEST_FILE_2_PATH, self.OUTPUT_BASE_FILE_PATH+'.xls')
        self.assertEqual(self.TESTS_DATA_1, self.matcher.input_data_1)
        self.assertEqual(self.TESTS_DATA_2, self.matcher.input_data_2)

    def test_two_xlsx_input_files(self):
        """Test readed data from two Excel files
        """
        self.matcher.start(self.TEST_FILE_1_PATH, self.TEST_FILE_2_PATH, self.OUTPUT_BASE_FILE_PATH+'.xlsx')
        self.assertEqual(self.TESTS_DATA_1, self.matcher.input_data_1)
        self.assertEqual(self.TESTS_DATA_2, self.matcher.input_data_2)

################################################################################
# Tests for full process                                                       #
################################################################################
    def test_process_without_matches(self):
        """Test 2 array without matches
        """
        argv = ['exec', self.TEST_FILE_1_PATH, '1', self.TEST_FILE_2_PATH, '1', self.OUTPUT_BASE_FILE_PATH+'.xlsx']
        result_argv, err = self.matcher.read_argv(argv)
        self.assertEqual(True, result_argv)
        self.matcher.start(self.TEST_FILE_1_PATH, self.TEST_FILE_2_PATH, self.OUTPUT_BASE_FILE_PATH+'.xlsx')
        self.assertEqual('Activated', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 1, 1))
        self.assertEqual('Money', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 1, 2))
        self.assertEqual('Name', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 1, 5))
        
    def test_process_with_matches(self):
        """Test 2 array without matches
        """
        argv = ['exec', self.TEST_FILE_1_PATH, '3', self.TEST_FILE_2_PATH, '2', self.OUTPUT_BASE_FILE_PATH+'.xlsx']
        result_argv, err = self.matcher.read_argv(argv)
        self.assertEqual(True, result_argv)
        self.matcher.start(self.TEST_FILE_1_PATH, self.TEST_FILE_2_PATH, self.OUTPUT_BASE_FILE_PATH+'.xlsx')
        self.assertEqual('Name', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 1, 1))
        self.assertEqual('Activated', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 1, 2))
        self.assertEqual('Favorite color', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 1, 6))
        self.assertEqual('Marc ASSIN', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 2, 1))
        self.assertEqual('London', self.get_cell_in_xlsx(self.OUTPUT_BASE_FILE_PATH+'.xlsx', 4, 5))

if __name__ == '__main__':
    unittest.main()
